# xlsx_reader.py
# 2025-11-07: Date-in-text support ("... от 22.09.2025 г.", "от «22» сентября 2025 г.")
#             + numeric formats with slashes, Russian month names, neighbor total detection,
#             + BIN context, trace, and config-driven aliases.
# 2025-11-12: reason: BIN context hardening — БИН извлекается только при подтверждённом контексте
#                   (рядом слова «бин/покупатель/поставщик» или алиасы из config.markers);
#                   строгая валидация БИН (12 цифр, не все одинаковые, не «000…»).
# 2025-11-12: reason: bugfix — BASE_0001.xlsx не находился БИН покупателя.
#                   Добавлен целевой проход по ярлыкам ("БИН покупателя", "ИИН/БИН покупателя",
#                   "БИН получателя/заказчика"), поиск значения справа/ниже, расширены маркеры
#                   и окно контекста; учитываются переносы на следующую строку/столбец.

import json, os, re
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime, timedelta
from openpyxl import load_workbook

# ---------------- util & config ----------------

def _project_root() -> str:
    return os.path.dirname(__file__)

def _load_config() -> Dict[str, Any]:
    candidates = [
        os.path.join(_project_root(), "config.json"),
        os.path.join(_project_root(), "ulyuly_checker", "config.json"),
        os.path.join(os.getcwd(), "config.json"),
    ]
    for p in candidates:
        try:
            if os.path.exists(p):
                with open(p, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, dict):
                    return data
        except Exception:
            pass
    return {}

_CFG = _load_config()

def _norm(s: Any) -> str:
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\u00A0", " ")
    s = re.sub(r"[\u200b\u200e\u202f\t\r\n]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _norm_key(s: Any) -> str:
    s = _norm(s).lower()
    s = s.replace("№", "номер")
    s = s.replace("иин/бин", "бин")
    s = s.replace("бин/иин", "бин")
    s = s.replace("счёт", "счет")
    s = s.replace("счет-фактура", "счет фактура")
    s = s.replace("получател", "покупател")  # 2025-11-12: reason: buyer markers normalize
    s = re.sub(r"[^\w\s]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _only_digits(s: str) -> str:
    return "".join(ch for ch in s if ch.isdigit())

def _is_valid_bin(value: Any) -> bool:
    """
    Жёсткая проверка БИН/ИИН:
      - только 12 цифр после очистки;
      - не «000000000000» и не «111111111111» (все одинаковые);
    """
    if value is None:
        return False
    try:
        s = _only_digits(str(value))
    except Exception:
        return False
    if len(s) != 12:
        return False
    if len(set(s)) == 1:
        return False
    return True

# ---------------- dates ----------------

_MONTHS_RU = {
    "января":1,"февраля":2,"марта":3,"апреля":4,"мая":5,"июня":6,
    "июля":7,"августа":8,"сентября":9,"октября":10,"ноября":11,"декабря":12
}

def _excel_serial_to_datetime(value: float) -> Optional[datetime]:
    try:
        base = datetime(1899, 12, 30)  # Windows base (with 1900 bug)
        return base + timedelta(days=float(value))
    except Exception:
        return None

def _as_excel_or_text_date(val: Any) -> Optional[datetime]:
    # 1) real datetime
    if isinstance(val, datetime):
        return val.replace(tzinfo=None)

    # 2) Excel serial number
    if isinstance(val, (int, float)):
        if 20000 <= float(val) <= 60000:
            dt = _excel_serial_to_datetime(val)
            if dt:
                return dt

    # 3) text with numeric or rus-month formats (optionally with "от ... г.")
    text = _norm(val)
    if not text:
        return None

    num_patterns = [
        r"(?P<d>\d{1,2})\.(?P<m>\d{1,2})\.(?P<y>\d{4})",
        r"(?P<y>\d{4})-(?P<m>\d{1,2})-(?P<d>\d{1,2})",
        r"(?P<d>\d{1,2})/(?P<m>\d{1,2})/(?P<y>\d{4})",
        r"(?P<y>\d{4})/(?P<m>\d{1,2})/(?P<d>\d{1,2})",
    ]
    for p in num_patterns:
        m = re.search(p, text)
        if m:
            try:
                y = int(m.group("y")); mth = int(m.group("m")); d = int(m.group("d"))
                return datetime(y, mth, d)
            except Exception:
                pass

    m = re.search(
        r"(?:\bот\s+)?[«\"]?(?P<d>\d{1,2})[»\"]?\s+(?P<mon>января|февраля|марта|апреля|мая|июня|июля|августа|сентября|октября|ноября|декабря)\s+(?P<y>\d{4})\s*(?:г\.?|г)?",
        text.lower()
    )
    if m:
        try:
            d = int(m.group("d"))
            y = int(m.group("y"))
            mth = _MONTHS_RU[m.group("mon")]
            return datetime(y, mth, d)
        except Exception:
            return None
    return None

def _to_iso_date(dt: Optional[datetime]) -> str:
    return dt.strftime("%Y-%m-%d") if isinstance(dt, datetime) else ""

# ---------------- sheet scan ----------------

def _collect_cells(ws) -> List[Tuple[int,int,Any]]:
    out = []
    for r in ws.iter_rows():
        for c in r:
            out.append((c.row, c.column, c.value))
    return out

def _cell(cells, r, c):
    return next((v2 for (rr,cc,v2) in cells if rr == r and cc == c), None)

def _near_text(cells: List[Tuple[int,int,Any]], row: int, col: int, radius: int=2) -> str:
    out = []
    for (r, c, v) in cells:
        if abs(r-row) <= radius and abs(c-col) <= radius:
            t = _norm(v)
            if t:
                out.append(t.lower())
    return " ".join(out)

# ---------------- totals & dates detection ----------------

def _find_number_near(cells, r, c, search_right=5, search_down=3) -> Optional[Tuple[float,int,int]]:
    # in same cell
    val = _cell(cells, r, c)
    txt = _norm(val)
    m = re.search(r"(-?\d[\d\s,\.]*)", txt.lower())
    if m:
        raw = m.group(1).replace(" ", "").replace(",", ".")
        try: return float(raw), r, c
        except: pass
    # right
    for dc in range(1, search_right+1):
        neigh = _cell(cells, r, c+dc)
        if neigh is None: continue
        if isinstance(neigh, (int, float)):
            try: return float(neigh), r, c+dc
            except: pass
        t = _norm(neigh)
        m2 = re.search(r"(-?\d[\d\s,\.]*)", t.lower())
        if m2:
            raw = m2.group(1).replace(" ", "").replace(",", ".")
            try: return float(raw), r, c+dc
            except: pass
    # down
    for dr in range(1, search_down+1):
        neigh = _cell(cells, r+dr, c)
        if neigh is None: continue
        if isinstance(neigh, (int, float)):
            try: return float(neigh), r+dr, c
            except: pass
        t = _norm(neigh)
        m2 = re.search(r"(-?\d[\d\s,\.]*)", t.lower())
        if m2:
            raw = m2.group(1).replace(" ", "").replace(",", ".")
            try: return float(raw), r+dr, c
            except: pass
    return None

def _detect_dates_and_totals(cells, aliases) -> Tuple[str,str,Dict[str,Any],Dict[str,str]]:
    thr = float(_CFG.get("header_fuzzy_threshold", 0.82))
    thr_date = min(thr, 0.75)

    issue_alias = aliases.get("issue_date", [])
    turn_alias  = aliases.get("turnover_date", [])
    net_alias   = aliases.get("total_net", [])
    vat_alias   = aliases.get("total_vat", [])
    gross_alias = aliases.get("total_gross", [])
    total_alias = aliases.get("total_amount", [])

    issue_candidates: List[Tuple[datetime,int,int]] = []
    turn_candidates:  List[Tuple[datetime,int,int]] = []
    totals: Dict[str, Tuple[float,int,int]] = {}

    def _best(header: str, alias_list: List[str]) -> float:
        h = _norm_key(header)
        best = 0.0
        for a in alias_list:
            a_norm = _norm_key(a)
            max_len = max(len(h), len(a_norm), 1)
            same = 0
            for ch1, ch2 in zip(h, a_norm):
                if ch1 == ch2:
                    same += 1
            ratio = same / max_len
            if h == a_norm: ratio = 1.0
            if ratio > best: best = ratio
        return best

    for (r,c,v) in cells:
        # --- Dates ---
        dt = _as_excel_or_text_date(v)
        if dt:
            left = _norm(_cell(cells, r, c-1))
            up   = _norm(_cell(cells, r-1, c))
            left_hit_issue = _best(left, issue_alias) >= thr_date
            up_hit_issue   = _best(up,   issue_alias) >= thr_date
            left_hit_turn  = _best(left, turn_alias)  >= thr_date
            up_hit_turn    = _best(up,   turn_alias)  >= thr_date

            if left_hit_issue or up_hit_issue:
                issue_candidates.append((dt, r, c))
            if left_hit_turn or up_hit_turn:
                turn_candidates.append((dt, r, c))

            same_text = _norm(_cell(cells, r, c))
            same_key  = _norm_key(same_text)
            if not (left_hit_issue or up_hit_issue or left_hit_turn or up_hit_turn):
                if _best(same_text, issue_alias) >= thr_date or any(k in same_key for k in ("счет", "сф", "invoice")):
                    issue_candidates.append((dt, r, c))

        # --- Totals ---
        t = _norm(v)
        if t:
            for kind, aliases_list in (("total_net", net_alias),
                                       ("total_vat", vat_alias),
                                       ("total_gross", gross_alias),
                                       ("total_amount", total_alias)):
                if _best(t, aliases_list) >= thr:
                    found = _find_number_near(cells, r, c)
                    if found:
                        totals[kind] = found

    issue_candidates.sort(key=lambda x: (x[1], x[2]))
    turn_candidates.sort(key=lambda x: (x[1], x[2]))

    issue_date = _to_iso_date(issue_candidates[0][0] if issue_candidates else None)
    turnover_date = _to_iso_date(turn_candidates[0][0] if turn_candidates else None)

    totals_out: Dict[str, Any] = {}
    trace: Dict[str, str] = {}

    if "total_net" in totals:
        v,r,c = totals["total_net"];  totals_out["total_net"] = v;  trace["total_net"] = f"TOTAL_NET@R{r}C{c}"
    if "total_vat" in totals:
        v,r,c = totals["total_vat"];  totals_out["total_vat"] = v;  trace["total_vat"] = f"TOTAL_VAT@R{r}C{c}"
    if "total_gross" in totals:
        v,r,c = totals["total_gross"]; totals_out["total_gross"] = v; trace["total_gross"] = f"TOTAL_GROSS@R{r}C{c}"

    prefer = str((_CFG.get("totals", {})).get("prefer_total", "gross")).lower()
    if prefer == "net" and "total_net" in totals_out:
        totals_out["total_amount"] = totals_out["total_net"]
    elif prefer == "vat" and "total_vat" in totals_out:
        totals_out["total_amount"] = totals_out["total_vat"]
    elif "total_gross" in totals_out:
        totals_out["total_amount"] = totals_out["total_gross"]
    elif "total_amount" in totals:
        v,r,c = totals["total_amount"]; totals_out["total_amount"] = v; trace["total_amount"] = f"TOTAL@R{r}C{c}"

    return issue_date, turnover_date, totals_out, trace

# ---------------- BIN detection ----------------

# 2025-11-12: reason: расширенные маркеры для покупателя/поставщика
_SUPPLIER_TOKENS = {
    "поставщик","поставщика","продавец","продавца","продавцу","продавцом",
    "supplier","seller"
}
_BUYER_TOKENS = {
    "покупатель","покупателя","покупателю","покупателем",
    "получатель","получателя","заказчик","заказчика",
    "buyer","customer","recipient"
}
_BIN_TOKENS = {"бин","иин","iin","bin"}

def _has_any_token(key: str, tokens: set) -> bool:
    k = _norm_key(key)
    return any(tok in k for tok in tokens)

def _search_value_right_down(cells, r, c, right=6, down=4) -> Optional[Tuple[str,int,int]]:
    """Ищем 12-значное значение справа/ниже от ярлыка — учитывает типичные табличные макеты."""
    # same cell
    v = _cell(cells, r, c)
    digits = _only_digits(_norm(v))
    if _is_valid_bin(digits):
        return digits, r, c
    # right cells
    for dc in range(1, right+1):
        v = _cell(cells, r, c+dc)
        if v is None: continue
        digits = _only_digits(_norm(v))
        if _is_valid_bin(digits):
            return digits, r, c+dc
    # below cells
    for dr in range(1, down+1):
        v = _cell(cells, r+dr, c)
        if v is None: continue
        digits = _only_digits(_norm(v))
        if _is_valid_bin(digits):
            return digits, r+dr, c
    return None

def _detect_bins(cells, markers_cfg) -> Tuple[str,str,Dict[str,str]]:
    """
    Стратегия:
      A) Целевой проход: находим ярлыки, содержащие BIN/ИИН + «покупател/получател/заказчик» или «поставщик/продавец»,
         и берём ближайшие 12 цифр справа/ниже (устойчиво к разрыву по строкам/столбцам).
      B) Если не найдено — контекстное окно (старый подход), но с расширенным радиусом и маркерами.
    """
    supplier_bin, buyer_bin = "", ""
    trace: Dict[str, str] = {}

    # --- A) label-based pass ---
    for (r,c,v) in cells:
        text = _norm(v)
        if not text:
            continue
        key = _norm_key(text)
        if not _has_any_token(key, _BIN_TOKENS):
            continue

        # Поставщик
        if _has_any_token(key, _SUPPLIER_TOKENS):
            found = _search_value_right_down(cells, r, c, right=8, down=4)
            if found and not supplier_bin:
                supplier_bin = found[0]
                trace["supplier_bin"] = f"LABEL@R{r}C{c}->R{found[1]}C{found[2]}"

        # Покупатель
        if _has_any_token(key, _BUYER_TOKENS):
            found = _search_value_right_down(cells, r, c, right=8, down=4)
            if found and not buyer_bin:
                buyer_bin = found[0]
                trace["buyer_bin"] = f"LABEL@R{r}C{c}->R{found[1]}C{found[2]}"

    # --- B) fallback: context window (если не нашли на ярлыках) ---
    if not supplier_bin or not buyer_bin:
        for (r,c,v) in cells:
            t_raw = _norm(v)
            if not t_raw:
                continue
            digits = _only_digits(t_raw)
            if not _is_valid_bin(digits):
                continue
            ctx = _near_text(cells, r, c, radius=3)  # 2025-11-12: radius 3 -> 4? оставим 3
            ctx_key = _norm_key(ctx + " " + t_raw)

            if not supplier_bin and _has_any_token(ctx_key, _SUPPLIER_TOKENS) and _has_any_token(ctx_key, _BIN_TOKENS):
                supplier_bin = digits
                trace["supplier_bin"] = f"CTX@R{r}C{c}"
                continue

            if not buyer_bin and _has_any_token(ctx_key, _BUYER_TOKENS) and _has_any_token(ctx_key, _BIN_TOKENS):
                buyer_bin = digits
                trace["buyer_bin"] = f"CTX@R{r}C{c}"
                continue

    return supplier_bin, buyer_bin, trace

# ---------------- public API ----------------

def read_xlsx(file_path: str) -> Dict[str, Any]:
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    cells = _collect_cells(ws)

    aliases = _CFG.get("aliases", {})

    supplier_bin, buyer_bin, trace_bins = _detect_bins(cells, _CFG.get("markers", {}))
    issue_date, turnover_date, totals, trace_dt = _detect_dates_and_totals(cells, aliases)

    content: Dict[str, Any] = {
        "supplier_bin": supplier_bin or "",
        "buyer_bin": buyer_bin or "",
        "issue_date": issue_date or "",
        "turnover_date": turnover_date or "",
        "total_amount": totals.get("total_amount", ""),
        "total_net": totals.get("total_net", ""),
        "total_vat": totals.get("total_vat", ""),
        "total_gross": totals.get("total_gross", ""),
        "lines": [],
        "_trace": {**trace_bins, **trace_dt}
    }
    return content

def extract_data(file_path: str) -> Dict[str, Any]:
    return read_xlsx(file_path)

if __name__ == "__main__":
    import sys, json as _json
    if len(sys.argv) < 2:
        print("Usage: python xlsx_reader.py <file.xlsx>")
        raise SystemExit(1)
    data = read_xlsx(sys.argv[1])
    print(_json.dumps(data, ensure_ascii=False, indent=2))
