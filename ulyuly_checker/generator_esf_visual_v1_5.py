#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generator_esf_visual_v2_5.py — ULYULYU CHECKER Synthetic Dataset Generator
Дата: 2025-11-09
Автор: frukt22

Версия: 2.5 — BIN/DATE/AMOUNT Mutations Pack
---------------------------------------------------------------------------
# 2025-11-09 (причина: расширить покрытие тестов; обеспечить различие БИНов,
#              варьировать даты и суммы; безопасная запись в мердж-ячейки)
# Изменения по сравнению с v2.4:
#  • БАЗА: гарантируем supplier_BIN != recipient_BIN
#  • ДОБАВЛЕНЫ НОВЫЕ МУТАЦИИ:
#      BIN007 — одинаковые БИНы продавца и покупателя (нельзя)
#      D003   — слишком старая дата (до 2000-01-01)
#      D004   — дата в формате DD.MM.YYYY (не ISO)
#      TOT004 — расхождение по округлению (+0.01 к итогу)
#      AMT001 — total_amount как строка ("ABC")
#  • SAFE Excel write: запись в левую-верхнюю ячейку объединённого диапазона
#  • PDF: шрифт Arial с uni=True (кириллица)
#  • Обновлён словарь CHECKS
# Зависимости: openpyxl, fpdf2
---------------------------------------------------------------------------

Запуск (пример):
    python generator_esf_visual_v2_5.py ^
      --excel_layout "C:\\path\\to\\esf_form.xlsx" ^
      --template_json "C:\\path\\to\\assets\\templates\\esf_template.json"
"""
import argparse
import csv
import json
import os
import random
import string
from datetime import date, timedelta, datetime
from typing import Dict, Any

from fpdf import FPDF
from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple, get_column_letter

# === Пути по умолчанию (настрой через аргументы CLI) ===
TEMPLATE_JSON_PATH_DEFAULT = "assets/templates/esf_template.json"
EXCEL_LAYOUT_PATH_DEFAULT = "assets/templates/esf_form_v2019.xlsx"
OUTPUT_DIR_DEFAULT = "synthetic_esf_visual"
ARIAL_TTF = r"C:\Windows\Fonts\arial.ttf"


# ============================================================
# 🧩 Утилиты
# ============================================================
def rnd_round(value: float, ndigits: int = 2) -> float:
    sign = 1 if value >= 0 else -1
    factor = 10 ** ndigits
    return sign * (int(abs(value) * factor + 0.5)) / factor


def gen_BIN() -> str:
    return "".join(random.choices(string.digits, k=12))


def gen_date_iso(within_days: int = 60) -> str:
    d = date.today() - timedelta(days=random.randint(0, within_days))
    return d.strftime("%Y-%m-%d")


def gen_amount(min_v=1000, max_v=100000) -> float:
    return rnd_round(random.uniform(min_v, max_v))


def gen_signature() -> str:
    return "Подписано ЭЦП"


def load_esf_template(path: str) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data.get("esf_template", data)


# ============================================================
# 📊 Пересчёт итогов
# ============================================================
def recalc_totals(doc: Dict[str, Any]):
    """Пересчитывает total_amount, total_vat и total_with_vat на основе строк."""
    total_amount = sum(float(line.get("amount", 0) or 0) for line in doc.get("lines", []))
    total_vat = sum(float(line.get("vat_amount", 0) or 0) for line in doc.get("lines", []))
    total_with_vat = total_amount + total_vat
    doc["total_amount"] = rnd_round(total_amount, 2)
    doc["total_vat"] = rnd_round(total_vat, 2)
    doc["total_with_vat"] = rnd_round(total_with_vat, 2)


# ============================================================
# 🧠 Создание базового документа
# ============================================================
def make_base_document(doc_id: str, tmpl: Dict[str, Any]) -> Dict[str, Any]:
    doc = {"doc_id": doc_id}

    # Оригинальные ключи из шаблона
    for field in tmpl.get("key_fields", {}):
        lname = field.lower()
        if "supplier" in lname and "bin" in lname:
            doc[field] = gen_BIN()
        elif ("recipient" in lname or "buyer" in lname) and "bin" in lname:
            b = gen_BIN()
            # Гарантируем, что БИНы различны
            while b == doc.get("supplier_BIN"):
                b = gen_BIN()
            doc[field] = b
        elif "date" in lname:
            doc[field] = gen_date_iso()
        elif "total" in lname or "amount" in lname:
            doc[field] = gen_amount()
        elif "signature" in lname:
            doc[field] = gen_signature()
        else:
            doc[field] = ""

    # Примитивные строки номенклатуры
    doc["lines"] = [
        {"line_no": 1, "name": "Услуги связи",     "uom": "усл", "qty": 1, "price": 100000, "amount": 100000, "vat_rate": 12, "vat_amount": 12000},
        {"line_no": 2, "name": "Монтаж оборудования","uom": "усл", "qty": 1, "price":  50000, "amount":  50000, "vat_rate": 12, "vat_amount":  6000},
    ]
    recalc_totals(doc)
    return doc


# ============================================================
# 🧠 Мутации (старые + новые)
# ============================================================
# --- Суммы / Итого ---
def mutate_TOT001(doc):
    """Итог не совпадает со суммой строк: +10.00 к total_amount."""
    recalc_totals(doc)
    doc["total_amount"] = rnd_round(float(doc["total_amount"]) + 10.0, 2)

def mutate_TOT002(doc):
    """Итог с НДС меньше рассчитанного (-5%)."""
    recalc_totals(doc)
    doc["total_with_vat"] = rnd_round(float(doc["total_with_vat"]) * 0.95, 2)

def mutate_TOT003(doc):
    """Итог = 0 (все суммы нулевые)."""
    doc["total_amount"] = 0
    doc["total_vat"] = 0
    doc["total_with_vat"] = 0

def mutate_TOT004(doc):
    """Округление: +0.01 к итогу без НДС (расхождение по копейке)."""
    recalc_totals(doc)
    doc["total_amount"] = rnd_round(float(doc["total_amount"]) + 0.01, 2)

def mutate_NEG001(doc):
    """Отрицательные суммы строк."""
    for line in doc.get("lines", []):
        line["amount"] = -abs(float(line.get("amount", 0) or 0))
    recalc_totals(doc)

def mutate_NEG002(doc):
    """Отрицательные суммы НДС по строкам."""
    for line in doc.get("lines", []):
        line["vat_amount"] = -abs(float(line.get("vat_amount", 0) or 0))
    recalc_totals(doc)

def mutate_AMT001(doc):
    """total_amount задан нечисловой строкой."""
    recalc_totals(doc)
    doc["total_amount"] = "ABC"

# --- БИНы ---
def mutate_ID001(doc):  # supplier формат неверный
    doc["supplier_BIN"] = "12345"

def mutate_ID002(doc):  # recipient формат неверный
    doc["recipient_BIN"] = "12345"

def mutate_BIN001(doc):  # supplier пустой
    doc["supplier_BIN"] = ""

def mutate_BIN2(doc):  # историческое имя — не используем (оставлено для совместимости)
    doc["recipient_BIN"] = ""

def mutate_BIN002(doc):  # recipient пустой
    doc["recipient_BIN"] = ""

def mutate_BIN003(doc):  # supplier длина неверная
    doc["supplier_BIN"] = "1234567"

def mutate_BIN004(doc):  # recipient содержит буквы
    doc["recipient_BIN"] = "12AB56789012"

def mutate_BIN005(doc):  # recipient длина неверная (короче)
    doc["recipient_BIN"] = "98765"

def mutate_BIN006(doc):  # supplier содержит буквы
    doc["supplier_BIN"] = "AB1234567890"

def mutate_BIN007(doc):  # НОВОЕ: БИНы продавца и покупателя одинаковые
    b = doc.get("supplier_BIN") or gen_BIN()
    doc["supplier_BIN"] = b
    doc["recipient_BIN"] = b

# --- Даты ---
def mutate_D000(doc):  # дата отсутствует
    doc["date_issue"] = ""

def mutate_D001(doc):  # будущая дата
    doc["date_issue"] = (date.today() + timedelta(days=5)).strftime("%Y-%m-%d")

def mutate_D002(doc):  # некорректный ISO формат
    doc["date_issue"] = "2025/13/99"

def mutate_D004(doc):  # НОВОЕ: альтернативный формат DD.MM.YYYY
    doc["date_issue"] = datetime.today().strftime("%d.%m.%Y")

def mutate_D003(doc):  # НОВОЕ: слишком старая дата (для сценария «не в допустимом периоде»)
    doc["date_issue"] = "1998-01-01"


# Справочник доступных мутаций
MUTATIONS = {
    # суммы
    "TOT001": mutate_TOT001,
    "TOT002": mutate_TOT002,
    "TOT003": mutate_TOT003,
    "TOT004": mutate_TOT004,   # новое
    "NEG001": mutate_NEG001,
    "NEG002": mutate_NEG002,
    "AMT001": mutate_AMT001,   # новое

    # BIN / ID
    "ID001":  mutate_ID001,
    "ID002":  mutate_ID002,
    "BIN001": mutate_BIN001,
    "BIN002": mutate_BIN002,
    "BIN003": mutate_BIN003,
    "BIN004": mutate_BIN004,
    "BIN005": mutate_BIN005,
    "BIN006": mutate_BIN006,
    "BIN007": mutate_BIN007,   # новое

    # даты
    "D000": mutate_D000,
    "D001": mutate_D001,
    "D002": mutate_D002,
    "D003": mutate_D003,       # новое
    "D004": mutate_D004,       # новое
}

# Ожидаемые статусы (для truth.csv)
CHECKS = {
    # суммы
    "TOT001": ("TOT001", "ERROR",   "Сумма не совпадает с суммой строк."),
    "TOT002": ("TOT002", "ERROR",   "Итог меньше рассчитанного."),
    "TOT003": ("TOT003", "WARNING", "Сумма равна нулю."),
    "TOT004": ("TOT004", "WARNING", "Расхождение по округлению."),
    "NEG001": ("NEG001", "ERROR",   "Строки содержат отрицательные суммы."),
    "NEG002": ("NEG002", "ERROR",   "НДС не может быть отрицательным."),
    "AMT001": ("AMT001", "ERROR",   "Итоговая сумма имеет нечисловой формат."),

    # BIN / ID
    "ID001":  ("ID001",  "ERROR", "БИН поставщика имеет неверный формат."),
    "ID002":  ("ID002",  "ERROR", "БИН покупателя имеет неверный формат."),
    "BIN001": ("BIN001", "ERROR", "БИН поставщика отсутствует."),
    "BIN002": ("BIN002", "ERROR", "БИН покупателя отсутствует."),
    "BIN003": ("BIN003", "ERROR", "БИН поставщика имеет неверную длину."),
    "BIN004": ("BIN004", "ERROR", "БИН покупателя содержит недопустимые символы."),
    "BIN005": ("BIN005", "ERROR", "БИН покупателя имеет неверную длину."),
    "BIN006": ("BIN006", "ERROR", "БИН поставщика содержит недопустимые символы."),
    "BIN007": ("BIN007", "ERROR", "БИНы продавца и покупателя совпадают."),

    # даты
    "D000": ("D000", "ERROR",   "Дата выставления отсутствует."),
    "D001": ("D001", "ERROR",   "Дата ЭСФ не может быть в будущем."),
    "D002": ("D002", "ERROR",   "Дата имеет некорректный формат."),
    "D003": ("D003", "WARNING", "Дата слишком старая для текущего периода."),
    "D004": ("D004", "ERROR",   "Дата указана не в формате ГГГГ-ММ-ДД."),
}


# ============================================================
# 📊 Excel (заполнение layout) с поддержкой merged cells
# ============================================================
EXCEL_POSITIONS = {
    "date_issue":   "B5",
    "supplier_BIN": "B10",
    "recipient_BIN":"B17",
    "total_amount": "B41",
    # таблица строк
    "lines_start_row": 28,
    "lines_cols": {  # настроить под свой бланк
        "line_no":   "A",
        "name":      "B",
        "qty":       "F",
        "price":     "G",
        "amount":    "H",
        "vat_amount":"J",
    },
}

def resolve_merged_top_left(ws, coord: str) -> str:
    """Если coord в объединённом диапазоне — вернуть адрес его top-left ячейки."""
    r, c = coordinate_to_tuple(coord)
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= r <= mr.max_row and mr.min_col <= c <= mr.max_col:
            return f"{get_column_letter(mr.min_col)}{mr.min_row}"
    return coord

def safe_set_cell(ws, coord: str, value):
    ws[resolve_merged_top_left(ws, coord)] = value

def write_excel_from_layout(doc: Dict[str, Any], layout_path: str, output_path: str):
    wb = load_workbook(layout_path)
    ws = wb.active

    # простые поля
    mapping = [
        ("date_issue",   EXCEL_POSITIONS["date_issue"]),
        ("supplier_BIN", EXCEL_POSITIONS["supplier_BIN"]),
        ("recipient_BIN",EXCEL_POSITIONS["recipient_BIN"]),
        ("total_amount", EXCEL_POSITIONS["total_amount"]),
    ]
    for key, cell in mapping:
        if key in doc:
            safe_set_cell(ws, cell, doc[key])

    # строки
    start_row = EXCEL_POSITIONS["lines_start_row"]
    cols = EXCEL_POSITIONS["lines_cols"]
    for i, line in enumerate(doc.get("lines", [])):
        r = start_row + i
        safe_set_cell(ws, f"{cols['line_no']}{r}",    line.get("line_no"))
        safe_set_cell(ws, f"{cols['name']}{r}",       line.get("name"))
        safe_set_cell(ws, f"{cols['qty']}{r}",        line.get("qty"))
        safe_set_cell(ws, f"{cols['price']}{r}",      line.get("price"))
        safe_set_cell(ws, f"{cols['amount']}{r}",     line.get("amount"))
        safe_set_cell(ws, f"{cols['vat_amount']}{r}", line.get("vat_amount"))

    wb.save(output_path)


# ============================================================
# 🖨️ PDF (визуальный макет)
# ============================================================
def write_pdf_visual(doc: Dict[str, Any], output_path: str):
    if not os.path.exists(ARIAL_TTF):
        raise FileNotFoundError("Не найден Arial.ttf (C:\\Windows\\Fonts). Установите системный шрифт.")

    pdf = FPDF(unit="mm", format="A4")
    pdf.add_page()
    pdf.add_font("Arial", "", ARIAL_TTF, uni=True)
    pdf.add_font("Arial", "B", ARIAL_TTF, uni=True)

    pdf.set_font("Arial", "", 12)
    pdf.cell(0, 8, "ЭЛЕКТРОННЫЙ СЧЁТ-ФАКТУРА", ln=1, align="C")
    pdf.set_font("Arial", "", 9)
    pdf.cell(0, 6, f"Номер документа: {doc.get('doc_id','')}", ln=1, align="C")
    pdf.cell(0, 6, f"Дата выписки: {doc.get('date_issue','')}", ln=1, align="C")
    pdf.ln(4)

    pdf.set_font("Arial", "B", 10)
    pdf.cell(0, 6, "Раздел B. Поставщик", ln=1)
    pdf.set_font("Arial", "", 9)
    pdf.cell(0, 5, f"БИН поставщика: {doc.get('supplier_BIN','')}", ln=1)
    pdf.ln(2)

    pdf.set_font("Arial", "B", 10)
    pdf.cell(0, 6, "Раздел C. Получатель", ln=1)
    pdf.set_font("Arial", "", 9)
    pdf.cell(0, 5, f"БИН покупателя: {doc.get('recipient_BIN','')}", ln=1)
    pdf.ln(4)

    # Таблица
    pdf.set_font("Arial", "B", 9)
    headers = ["№", "Наименование", "Кол-во", "Цена", "Сумма", "НДС"]
    widths =  [10, 65, 15, 25, 25, 25]
    for h, w in zip(headers, widths):
        pdf.cell(w, 7, h, 1, align="C")
    pdf.ln()
    pdf.set_font("Arial", "", 8)
    for line in doc.get("lines", []):
        pdf.cell(10, 6, str(line.get("line_no","")), 1)
        pdf.cell(65, 6, str(line.get("name","")), 1)
        pdf.cell(15, 6, str(line.get("qty","")), 1)
        pdf.cell(25, 6, str(line.get("price","")), 1)
        pdf.cell(25, 6, str(line.get("amount","")), 1)
        pdf.cell(25, 6, str(line.get("vat_amount","")), 1, ln=1)

    pdf.ln(4)
    pdf.set_font("Arial", "B", 10)
    pdf.cell(0, 6, f"Всего без НДС: {doc.get('total_amount','')}", ln=1)
    pdf.cell(0, 6, f"Сумма НДС: {doc.get('total_vat','')}", ln=1)
    pdf.cell(0, 6, f"Всего с НДС: {doc.get('total_with_vat','')}", ln=1)
    pdf.ln(6)
    pdf.set_font("Arial", "", 9)
    pdf.cell(0, 6, f"Подпись: {doc.get('signature_ECP','')}", ln=1)

    pdf.output(output_path)


# ============================================================
# 🚀 Генерация датасета
# ============================================================
def generate_dataset_visual(out_dir, excel_layout_path, template_json_path,
                            n_base, mutations_per_base, seed=42):
    random.seed(seed)
    tmpl = load_esf_template(template_json_path)

    os.makedirs(out_dir, exist_ok=True)
    docs_dir = os.path.join(out_dir, "invoices")
    os.makedirs(docs_dir, exist_ok=True)
    truth_path = os.path.join(out_dir, "truth.csv")

    with open(truth_path, "w", newline="", encoding="utf-8") as tf:
        writer = csv.writer(tf)
        writer.writerow(["doc_id", "check_code", "expected_status", "expected_message"])

        # Базовые документы (корректные, с разными БИНами)
        for i in range(n_base):
            base_id = f"BASE_{i+1:04d}"
            doc = make_base_document(base_id, tmpl)
            write_excel_from_layout(doc, excel_layout_path, os.path.join(docs_dir, f"{base_id}.xlsx"))
            write_pdf_visual(doc, os.path.join(docs_dir, f"{base_id}.pdf"))
            writer.writerow([base_id, "BASE", "OK", "Базовый корректный документ."])

        # Мутированные документы
        codes = list(MUTATIONS.keys())
        for i in range(n_base):
            base = make_base_document(f"TEMPLATE_{i+1:04d}", tmpl)
            chosen = random.sample(codes, k=min(mutations_per_base, len(codes)))
            for code in chosen:
                doc_id = f"MUT_{i+1:04d}_{code}"
                doc = json.loads(json.dumps(base))
                doc["doc_id"] = doc_id
                MUTATIONS[code](doc)

                write_excel_from_layout(doc, excel_layout_path, os.path.join(docs_dir, f"{doc_id}.xlsx"))
                write_pdf_visual(doc, os.path.join(docs_dir, f"{doc_id}.pdf"))

                chk = CHECKS[code]
                writer.writerow([doc_id, chk[0], chk[1], chk[2]])

    print(f"✅ Визуальный датасет создан: {out_dir}")
    print(f" - Документы: {docs_dir}")
    print(f" - Истина:    {truth_path}")


# ============================================================
# CLI
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="ULYULYU Visual ESF Generator v2.5 — BIN/DATE/AMOUNT Mutations Pack")
    parser.add_argument("--out_dir", type=str, default=OUTPUT_DIR_DEFAULT)
    parser.add_argument("--excel_layout", type=str, default=EXCEL_LAYOUT_PATH_DEFAULT)
    parser.add_argument("--template_json", type=str, default=TEMPLATE_JSON_PATH_DEFAULT)
    parser.add_argument("--n_base", type=int, default=3)
    parser.add_argument("--mutations_per_base", type=int, default=8)
    parser.add_argument("--seed", type=int, default=42)
    args = parser.parse_args()

    if not os.path.exists(args.excel_layout):
        raise FileNotFoundError(f"Excel-бланк не найден: {args.excel_layout}")
    if not os.path.exists(args.template_json):
        raise FileNotFoundError(f"Шаблон не найден: {args.template_json}")
    if not os.path.exists(ARIAL_TTF):
        raise FileNotFoundError("Не найден Arial.ttf (C:\\Windows\\Fonts). Установите системный шрифт.")

    generate_dataset_visual(args.out_dir, args.excel_layout, args.template_json,
                            args.n_base, args.mutations_per_base, args.seed)


if __name__ == "__main__":
    main()
