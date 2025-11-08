#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generator_esf_visual_v2_4.py — ULYULYU CHECKER Synthetic Dataset Generator
Дата: 2025-11-09
Автор: frukt22

Версия: 2.4 — Symmetric BIN Edition
-------------------------------------------------------------
Изменения относительно v1.5:
✅ Добавлены ID002, BIN005, BIN006 (зеркальные ошибки по БИН)
✅ Обновлён словарь CHECKS
✅ Пересчёт итогов сохранён
-------------------------------------------------------------
"""

import argparse
import csv
import json
import os
import random
import string
from datetime import date, timedelta
from typing import Dict, Any
from fpdf import FPDF
from fpdf.enums import XPos, YPos
from openpyxl import load_workbook

# === Пути ===
TEMPLATE_JSON_PATH_DEFAULT = "assets/templates/esf_template.json"
EXCEL_LAYOUT_PATH_DEFAULT = "assets/templates/esf_form_v2019.xlsx"
OUTPUT_DIR_DEFAULT = "synthetic_esf_visual"
ARIAL_TTF = r"C:\Windows\Fonts\arial.ttf"

# ============================================================
# 🧩 Утилиты
# ============================================================
def rnd_round(value: float, ndigits: int = 2) -> float:
    sign = 1 if value >= 0 else -1
    factor = 10 ** ndigits
    return sign * (int(abs(value) * factor + 0.5)) / factor

def gen_BIN() -> str:
    return "".join(random.choices(string.digits, k=12))

def gen_date_iso(within_days: int = 60) -> str:
    d = date.today() - timedelta(days=random.randint(0, within_days))
    return d.strftime("%Y-%m-%d")

def gen_amount(min_v=1000, max_v=100000) -> float:
    return rnd_round(random.uniform(min_v, max_v))

def gen_signature() -> str:
    return "Подписано ЭЦП"

def load_esf_template(path: str) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data.get("esf_template", data)

# ============================================================
# 📊 Пересчёт итогов
# ============================================================
def recalc_totals(doc):
    """Пересчитывает total_amount, total_vat и total_with_vat."""
    total_amount = sum(line.get("amount", 0) for line in doc.get("lines", []))
    total_vat = sum(line.get("vat_amount", 0) for line in doc.get("lines", []))
    total_with_vat = total_amount + total_vat
    doc["total_amount"] = round(total_amount, 2)
    doc["total_vat"] = round(total_vat, 2)
    doc["total_with_vat"] = round(total_with_vat, 2)

# ============================================================
# 🧠 Создание базового документа
# ============================================================
def make_base_document(doc_id: str, tmpl: Dict[str, Any]) -> Dict[str, Any]:
    doc = {"doc_id": doc_id}
    for field in tmpl.get("key_fields", {}):
        lname = field.lower()
        if "supplier" in lname and "bin" in lname:
            doc[field] = gen_BIN()
        elif ("recipient" in lname or "buyer" in lname) and "bin" in lname:
            b = gen_BIN()
            if b == doc.get("supplier_BIN"):
                b = gen_BIN()
            doc[field] = b
        elif "date" in lname:
            doc[field] = gen_date_iso()
        elif "total" in lname or "amount" in lname:
            doc[field] = gen_amount()
        elif "signature" in lname:
            doc[field] = gen_signature()
        else:
            doc[field] = ""
    doc["lines"] = [
        {"line_no": 1, "name": "Услуги связи", "uom": "усл", "qty": 1,
         "price": 100000, "amount": 100000, "vat_rate": "12%", "vat_amount": 12000},
        {"line_no": 2, "name": "Монтаж оборудования", "uom": "усл", "qty": 1,
         "price": 50000, "amount": 50000, "vat_rate": "12%", "vat_amount": 6000},
    ]
    recalc_totals(doc)
    return doc

# ============================================================
# 🧠 Мутации
# ============================================================
def mutate_TOT001(doc):
    recalc_totals(doc)
    doc["total_amount"] = rnd_round(float(doc["total_amount"]) + 10.0)

def mutate_TOT002(doc):
    """Ошибка: итог меньше рассчитанного (-5%)."""
    recalc_totals(doc)
    doc["total_with_vat"] = round(doc["total_with_vat"] * 0.95, 2)

def mutate_TOT003(doc):
    """Ошибка: сумма равна нулю."""
    doc["total_amount"] = 0
    doc["total_vat"] = 0
    doc["total_with_vat"] = 0

def mutate_NEG001(doc):
    """Ошибка: отрицательные суммы строк."""
    for line in doc.get("lines", []):
        line["amount"] = -abs(line.get("amount", 0))
    recalc_totals(doc)

def mutate_NEG002(doc):
    """Ошибка: отрицательный НДС."""
    for line in doc.get("lines", []):
        line["vat_amount"] = -abs(line.get("vat_amount", 0))
    recalc_totals(doc)

# --- BIN / ID ошибки ---
def mutate_ID001(doc): doc["supplier_BIN"] = "12345"
def mutate_ID002(doc): doc["recipient_BIN"] = "12345"
def mutate_BIN001(doc): doc["supplier_BIN"] = ""
def mutate_BIN002(doc): doc["recipient_BIN"] = ""
def mutate_BIN003(doc): doc["supplier_BIN"] = "1234567"
def mutate_BIN004(doc): doc["recipient_BIN"] = "12AB56789012"
def mutate_BIN005(doc): doc["recipient_BIN"] = "98765"
def mutate_BIN006(doc): doc["supplier_BIN"] = "AB1234567890"

# --- Даты ---
def mutate_D000(doc): doc["date_issue"] = ""
def mutate_D001(doc): doc["date_issue"] = (date.today() + timedelta(days=5)).strftime("%Y-%m-%d")
def mutate_D002(doc): doc["date_issue"] = "2025/13/99"

# ============================================================
# 🧩 Словари мутаций и проверок
# ============================================================
MUTATIONS = {
    "TOT001": mutate_TOT001,
    "TOT002": mutate_TOT002,
    "TOT003": mutate_TOT003,
    "NEG001": mutate_NEG001,
    "NEG002": mutate_NEG002,
    "ID001": mutate_ID001,
    "ID002": mutate_ID002,
    "BIN001": mutate_BIN001,
    "BIN002": mutate_BIN002,
    "BIN003": mutate_BIN003,
    "BIN004": mutate_BIN004,
    "BIN005": mutate_BIN005,
    "BIN006": mutate_BIN006,
    "D000": mutate_D000,
    "D001": mutate_D001,
    "D002": mutate_D002,
}

CHECKS = {
    "TOT001": ("TOT001", "ERROR", "Сумма не совпадает с суммой строк."),
    "TOT002": ("TOT002", "ERROR", "Итог меньше рассчитанного."),
    "TOT003": ("TOT003", "WARNING", "Сумма равна нулю."),
    "NEG001": ("NEG001", "ERROR", "Строки содержат отрицательные суммы."),
    "NEG002": ("NEG002", "ERROR", "НДС не может быть отрицательным."),
    "ID001": ("ID001", "ERROR", "БИН поставщика имеет неверный формат."),
    "ID002": ("ID002", "ERROR", "БИН покупателя имеет неверный формат."),
    "BIN001": ("BIN001", "ERROR", "БИН поставщика отсутствует."),
    "BIN002": ("BIN002", "ERROR", "БИН покупателя отсутствует."),
    "BIN003": ("BIN003", "ERROR", "БИН поставщика имеет неверную длину."),
    "BIN004": ("BIN004", "ERROR", "БИН покупателя содержит недопустимые символы."),
    "BIN005": ("BIN005", "ERROR", "БИН покупателя имеет неверную длину."),
    "BIN006": ("BIN006", "ERROR", "БИН поставщика содержит недопустимые символы."),
    "D000": ("D000", "ERROR", "Дата выставления отсутствует."),
    "D001": ("D001", "ERROR", "Дата ЭСФ не может быть в будущем."),
    "D002": ("D002", "ERROR", "Дата имеет некорректный формат."),
}

# ============================================================
# 📊 Excel
# ============================================================
def write_excel_from_layout(doc: Dict[str, Any], layout_path: str, output_path: str):
    wb = load_workbook(layout_path)
    ws = wb.active
    ws["B5"] = doc.get("date_issue", "")
    ws["B10"] = doc.get("supplier_BIN", "")
    ws["B17"] = doc.get("recipient_BIN", "")
    ws["B41"] = doc.get("total_amount", "")
    row = 28
    for line in doc.get("lines", []):
        ws[f"A{row}"] = line.get("line_no", "")
        ws[f"B{row}"] = line.get("name", "")
        ws[f"F{row}"] = line.get("qty", "")
        ws[f"G{row}"] = line.get("price", "")
        ws[f"H{row}"] = line.get("amount", "")
        ws[f"J{row}"] = line.get("vat_amount", "")
        row += 1
    wb.save(output_path)

# ============================================================
# 🖨️ PDF
# ============================================================
def write_pdf_visual(doc: Dict[str, Any], output_path: str):
    pdf = FPDF(unit="mm", format="A4")
    pdf.add_page()
    pdf.add_font("Arial", "", ARIAL_TTF)
    pdf.add_font("Arial", "B", ARIAL_TTF)
    pdf.set_font("Arial", "", 11)
    pdf.cell(0, 8, "ЭЛЕКТРОННЫЙ СЧЁТ-ФАКТУРА", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
    pdf.set_font("Arial", "", 9)
    pdf.cell(0, 6, f"Номер документа: {doc.get('doc_id','')}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
    pdf.cell(0, 6, f"Дата выписки: {doc.get('date_issue','')}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
    pdf.ln(4)
    pdf.set_font("Arial", "B", 10)
    pdf.cell(0, 6, "Раздел B. Поставщик", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Arial", "", 9)
    pdf.cell(0, 5, f"БИН поставщика: {doc.get('supplier_BIN','')}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(4)
    pdf.set_font("Arial", "B", 10)
    pdf.cell(0, 6, "Раздел C. Получатель", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Arial", "", 9)
    pdf.cell(0, 5, f"БИН покупателя: {doc.get('recipient_BIN','')}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(6)
    pdf.set_font("Arial", "B", 9)
    headers = ["№", "Наименование", "Кол-во", "Цена", "Сумма", "НДС"]
    widths = [10, 65, 15, 25, 25, 25]
    for h, w in zip(headers, widths): pdf.cell(w, 7, h, 1, align="C")
    pdf.ln()
    pdf.set_font("Arial", "", 8)
    for line in doc.get("lines", []):
        pdf.cell(10, 6, str(line.get("line_no","")), 1)
        pdf.cell(65, 6, str(line.get("name","")), 1)
        pdf.cell(15, 6, str(line.get("qty","")), 1)
        pdf.cell(25, 6, str(line.get("price","")), 1)
        pdf.cell(25, 6, str(line.get("amount","")), 1)
        pdf.cell(25, 6, str(line.get("vat_amount","")), 1, ln=1)
    pdf.ln(4)
    pdf.set_font("Arial", "B", 10)
    pdf.cell(0, 6, f"Всего без НДС: {doc.get('total_amount','')}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.cell(0, 6, f"Сумма НДС: {doc.get('total_vat','')}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.cell(0, 6, f"Всего с НДС: {doc.get('total_with_vat','')}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(8)
    pdf.set_font("Arial", "", 9)
    pdf.cell(0, 6, f"Подпись: {doc.get('signature_ECP','')}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.output(output_path)

# ============================================================
# 🚀 Генерация датасета
# ============================================================
def generate_dataset_visual(out_dir, excel_layout_path, template_json_path,
                            n_base, mutations_per_base, seed=42):
    random.seed(seed)
    tmpl = load_esf_template(template_json_path)
    os.makedirs(out_dir, exist_ok=True)
    docs_dir = os.path.join(out_dir, "invoices")
    os.makedirs(docs_dir, exist_ok=True)
    truth_path = os.path.join(out_dir, "truth.csv")

    with open(truth_path, "w", newline="", encoding="utf-8") as tf:
        writer = csv.writer(tf)
        writer.writerow(["doc_id", "check_code", "expected_status", "expected_message"])
        for i in range(n_base):
            base_id = f"BASE_{i+1:04d}"
            doc = make_base_document(base_id, tmpl)
            write_excel_from_layout(doc, excel_layout_path, os.path.join(docs_dir, f"{base_id}.xlsx"))
            write_pdf_visual(doc, os.path.join(docs_dir, f"{base_id}.pdf"))
            writer.writerow([base_id, "BASE", "OK", "Базовый корректный документ."])
        codes = list(MUTATIONS.keys())
        for i in range(n_base):
            base = make_base_document(f"TEMPLATE_{i+1:04d}", tmpl)
            chosen = random.sample(codes, k=min(mutations_per_base, len(codes)))
            for code in chosen:
                doc_id = f"MUT_{i+1:04d}_{code}"
                doc = json.loads(json.dumps(base))
                doc["doc_id"] = doc_id
                MUTATIONS[code](doc)
                write_excel_from_layout(doc, excel_layout_path, os.path.join(docs_dir, f"{doc_id}.xlsx"))
                write_pdf_visual(doc, os.path.join(docs_dir, f"{doc_id}.pdf"))
                chk = CHECKS[code]
                writer.writerow([doc_id, chk[0], chk[1], chk[2]])

    print(f"✅ Визуальный датасет создан: {out_dir}")
    print(f" - Документы: {docs_dir}")
    print(f" - Истина: {truth_path}")

# ============================================================
# CLI
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="ULYULYU Visual ESF Generator v2.4 — Symmetric BIN Edition")
    parser.add_argument("--out_dir", type=str, default=OUTPUT_DIR_DEFAULT)
    parser.add_argument("--excel_layout", type=str, default=EXCEL_LAYOUT_PATH_DEFAULT)
    parser.add_argument("--template_json", type=str, default=TEMPLATE_JSON_PATH_DEFAULT)
    parser.add_argument("--n_base", type=int, default=3)
    parser.add_argument("--mutations_per_base", type=int, default=5)
    parser.add_argument("--seed", type=int, default=42)
    args = parser.parse_args()
    generate_dataset_visual(args.out_dir, args.excel_layout, args.template_json,
                            args.n_base, args.mutations_per_base, args.seed)

if __name__ == "__main__":
    main()
